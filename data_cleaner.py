#!/usr/bin/env python3
"""
Pflegeheim Data Cleaner — Ollama + Web search → cleaned contact data

Pipeline per row:
  1. Pick first row with cleaned=FALSE from postgres
  2. DuckDuckGo search "{name} {ort} Impressum Kontakt Telefon Email"
  3. Fetch existing impressum/website page text
  4. Send all of it to Ollama (JSON mode) for extraction
  5. Validate (phone format, email format, PLZ in address, ort match)
  6. Write back: telefon_clean, email_clean, …, quality (ok|suspect|empty), cleaner

Usage:
  python data_cleaner.py --import-only                              # just import CSV
  python data_cleaner.py --loop --limit 50 --xlsx out/run.xlsx      # 50 rows + export
  python data_cleaner.py --xlsx-only --xlsx out/all.xlsx            # export only
  python data_cleaner.py --xlsx-only --xlsx-cleaner langdock --xlsx out/lang.xlsx

Env vars (or .env):
  PGHOST / PGPORT / PGDATABASE / PGUSER / PGPASSWORD
  OLLAMA_HOST  (default: http://localhost:11434)
  OLLAMA_MODEL (default: qwen3.5:9b)
  CSV_PATH     (default: data/pflegeheime_nrw.csv)
"""

import argparse
import csv
import json
import os
import re
import time
import logging
from datetime import timezone
import requests
from bs4 import BeautifulSoup
from ddgs import DDGS
from dotenv import load_dotenv

load_dotenv()

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "qwen3.5:9b")
CSV_PATH = os.getenv("CSV_PATH", "data/pflegeheime_nrw.csv")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "de-DE,de;q=0.9",
}

DDG_SNIPPET_LIMIT = 6

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS pflegeheime (
    api_id              TEXT PRIMARY KEY,
    name                TEXT,
    ort                 TEXT,
    kreis               TEXT,
    lat                 REAL,
    lon                 REAL,
    website             TEXT,
    impressum_url       TEXT,
    telefon             TEXT,
    email               TEXT,
    adresse             TEXT,
    geschaeftsfuehrung  TEXT,
    einrichtungsleitung TEXT,
    status              TEXT,
    error_msg           TEXT,
    cleaned             INTEGER DEFAULT 0,
    cleaned_at          TEXT,
    telefon_clean       TEXT,
    email_clean         TEXT,
    adresse_clean       TEXT,
    geschaeftsfuehrung_clean    TEXT,
    einrichtungsleitung_clean   TEXT,
    clean_notes         TEXT,
    quality             TEXT,
    cleaner             TEXT
);
"""

ALTER_TABLE_SQL = """
ALTER TABLE pflegeheime ADD COLUMN IF NOT EXISTS quality TEXT;
ALTER TABLE pflegeheime ADD COLUMN IF NOT EXISTS cleaner TEXT;
"""

PHONE_RE = re.compile(r"^[\+\d][\d\s/().+-]{6,30}$")
EMAIL_RE = re.compile(r"^[\w.+-]+@[\w-]+(\.[\w-]+)+$")
PLZ_RE = re.compile(r"\b\d{5}\b")

NO_DATA_RE = re.compile(
    r"^(no\s*(clear|specific)?\s*(data|information|email|phone)?(\s*found)?|"
    r"not\s*(specified|available|found|known)|"
    r"nicht\s*(spezifiziert|verfügbar|bekannt|angegeben)|"
    r"unbekannt|n/?a|none|null|tbd|tba|"
    r"keine\s*(angaben|info|daten)|"
    r"\[email[\s_]*protected\])\b.*",
    re.IGNORECASE,
)

JUNK_RE = re.compile(
    r"\b(wird\s+(übernommen|ergänzt|bekannt|angegeben)|"
    r"noch\s+nicht\s+(bekannt|angegeben)|"
    r"name\s+einfügen|hier\s+könnte|beispiel|mustermann|"
    r"siehe\s+(homepage|impressum|website)|"
    r"\[email[\s_]*protected\]|"
    r"no\s+(specific\s+)?information(\s+found)?|"
    r"not\s+specified|nicht\s+spezifiziert(\s+im\s+vorliegenden)?)\b",
    re.IGNORECASE,
)

CORP_FALLBACK_EMAILS = {
    "info@korian.de", "info@caritas.de", "info@drk.de", "info@awo.de",
    "info@asb.de", "info@johanniter.de", "info@diakonie.de", "info@malteser.de",
}

MUNICH_PHONE_RE = re.compile(r"^\+?49?\s*\(?0?\)?\s*89[\s\-]")
NO_DATA = "No clear Data"

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------
def db_connect():
    from marktradar.db import db_connect as _c
    return _c()


def ensure_schema(conn):
    from marktradar.db import bootstrap
    bootstrap(conn._c if hasattr(conn, "_c") else conn)


def table_empty(conn) -> bool:
    with conn.cursor() as cur:
        cur.execute("SELECT 1 FROM pflegeheime LIMIT 1")
        return cur.fetchone() is None


def import_csv(conn, csv_path: str):
    """Bulk-import CSV into psql (skips existing api_ids)."""
    imported = 0
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        with conn.cursor() as cur:
            for row in reader:
                cur.execute(
                    """
                    INSERT INTO pflegeheime
                        (api_id, name, ort, kreis, lat, lon, website, impressum_url,
                         telefon, email, adresse, geschaeftsfuehrung, einrichtungsleitung,
                         status, error_msg)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT (api_id) DO NOTHING
                    """,
                    (
                        row.get("api_id"), row.get("name"), row.get("ort"),
                        row.get("kreis"),
                        float(row["lat"]) if row.get("lat") else None,
                        float(row["lon"]) if row.get("lon") else None,
                        row.get("website"), row.get("impressum_url"),
                        row.get("telefon"), row.get("email"), row.get("adresse"),
                        row.get("geschaeftsfuehrung"), row.get("einrichtungsleitung"),
                        row.get("status"), row.get("error_msg"),
                    ),
                )
                imported += 1
    conn.commit()
    log.info(f"Imported {imported} rows from {csv_path}")


def get_next_uncleaned(conn, for_update: bool = False) -> dict | None:
    """Pick the next uncleaned row.

    If for_update=True, lock the row with FOR UPDATE SKIP LOCKED so that
    concurrent workers don't collide. Caller MUST commit/rollback to release.
    """
    lock_clause = "FOR UPDATE SKIP LOCKED" if for_update else ""
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT api_id, name, ort, kreis, website, impressum_url,
                   telefon, email, adresse, geschaeftsfuehrung, einrichtungsleitung
            FROM pflegeheime
            WHERE cleaned = FALSE
            ORDER BY api_id
            LIMIT 1
            {lock_clause}
            """
        )
        row = cur.fetchone()
        if not row:
            return None
        cols = ["api_id", "name", "ort", "kreis", "website", "impressum_url",
                "telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung"]
        return dict(zip(cols, row))


def save_cleaned(conn, api_id: str, data: dict, cleaner: str = "ollama"):
    with conn.cursor() as cur:
        cur.execute(
            """
            UPDATE pflegeheime SET
                cleaned = TRUE,
                cleaned_at = NOW(),
                telefon_clean = %s,
                email_clean = %s,
                adresse_clean = %s,
                geschaeftsfuehrung_clean = %s,
                einrichtungsleitung_clean = %s,
                clean_notes = %s,
                quality = %s,
                cleaner = %s
            WHERE api_id = %s
            """,
            (
                data.get("telefon"), data.get("email"), data.get("adresse"),
                data.get("geschaeftsfuehrung"), data.get("einrichtungsleitung"),
                data.get("notes"), data.get("quality"), cleaner, api_id,
            ),
        )
    conn.commit()


def heuristic_clean(row: dict) -> dict | None:
    """If the *scraped* fields already pass validation, skip search+LLM entirely.

    Returns a cleaned-style dict if confident, otherwise None.
    Match criterion: telefon passes PHONE_RE AND email passes EMAIL_RE
    AND adresse contains a PLZ AND (if known) the ort.
    """
    tel = (row.get("telefon") or "").strip()
    email = (row.get("email") or "").strip()
    addr = (row.get("adresse") or "").strip()
    ort = row.get("ort") or ""

    if not (PHONE_RE.match(tel) and EMAIL_RE.match(email)):
        return None
    if not PLZ_RE.search(addr):
        return None
    if ort and ort.lower() not in addr.lower():
        return None

    gf = (row.get("geschaeftsfuehrung") or "").strip() or "No clear Data"
    el = (row.get("einrichtungsleitung") or "").strip() or "No clear Data"

    return {
        "telefon": tel,
        "email": email,
        "adresse": addr,
        "geschaeftsfuehrung": gf,
        "einrichtungsleitung": el,
        "notes": "Heuristic: scraped fields pass all validators (skipped search+LLM)",
    }


def _is_data(v: str) -> bool:
    """True if the field contains real content (not a 'no data' placeholder or junk)."""
    s = (v or "").strip()
    if not s:
        return False
    if NO_DATA_RE.match(s):
        return False
    if JUNK_RE.search(s):
        return False
    return True


def validate_cleaned(cleaned: dict, ort: str) -> dict:
    """Plausibility checks. Adds 'quality' (ok|suspect|empty) + appends issues to notes."""
    issues: list[str] = []

    # Defensive: Mistral occasionally returns nested objects/lists/numbers — coerce to str.
    for k in ["telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung", "notes"]:
        v = cleaned.get(k)
        if v is None:
            cleaned[k] = ""
        elif isinstance(v, (dict, list)):
            cleaned[k] = json.dumps(v, ensure_ascii=False)
        elif not isinstance(v, str):
            cleaned[k] = str(v)

    tel = cleaned.get("telefon", "").strip()
    email = cleaned.get("email", "").strip()
    addr = cleaned.get("adresse", "").strip()
    gf = cleaned.get("geschaeftsfuehrung", "").strip()
    el = cleaned.get("einrichtungsleitung", "").strip()

    # --- Telefon ---
    if _is_data(tel):
        if not PHONE_RE.match(tel):
            issues.append(f"Telefon-Format ungewöhnlich: {tel}")
        if MUNICH_PHONE_RE.match(tel) and ort and ort.lower() not in {"münchen", "munich"}:
            issues.append(f"Telefon-Vorwahl 089 (München) passt nicht zu Ort '{ort}'")

    # --- Email ---
    if _is_data(email):
        if not EMAIL_RE.match(email):
            issues.append(f"Email-Format ungewöhnlich: {email}")
        if email.lower() in CORP_FALLBACK_EMAILS:
            issues.append(f"Generische Konzern-Fallback-Email: {email}")

    # --- Adresse ---
    if _is_data(addr):
        if not PLZ_RE.search(addr):
            issues.append("Adresse ohne PLZ")
        if ort and ort.lower() not in addr.lower():
            issues.append(f"Adresse-Ort '{ort}' nicht gefunden")

    # --- Personen-Felder: Junk-Detection (auch wenn _is_data false → cleanup) ---
    for fname, val in (("geschaeftsfuehrung", gf), ("einrichtungsleitung", el)):
        if val and JUNK_RE.search(val):
            issues.append(f"{fname}: Junk-/Boilerplate-Pattern: {val[:60]}")
            cleaned[fname] = "No clear Data"

    has_any = any(_is_data(cleaned.get(f, "")) for f in
                  ["telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung"])

    if not has_any:
        cleaned["quality"] = "empty"
    elif issues:
        cleaned["quality"] = "suspect"
    else:
        cleaned["quality"] = "ok"

    if issues:
        existing = (cleaned.get("notes") or "").strip()
        prefix = (existing + " | ") if existing else ""
        cleaned["notes"] = prefix + "; ".join(issues)

    return cleaned


# ---------------------------------------------------------------------------
# Search context
# ---------------------------------------------------------------------------
def _fetch_page_text(url: str) -> str:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=12, allow_redirects=True)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "lxml")
        for tag in soup(["script", "style", "nav", "header", "footer"]):
            tag.decompose()
        return soup.get_text(separator=" ", strip=True)[:3000]
    except Exception:
        return ""


def gather_context(row: dict) -> str:
    name = row["name"]
    ort = row["ort"]
    parts = []

    # Existing scraped data as baseline
    parts.append(f"=== Scraped data ===")
    for field in ["website", "telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung"]:
        if row.get(field):
            parts.append(f"{field}: {row[field]}")

    # Fresh web search
    query = f"{name} {ort} Impressum Kontakt Telefon Email"
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=DDG_SNIPPET_LIMIT, region="de-de"))
        parts.append(f"\n=== Web search: {query} ===")
        for r in results:
            parts.append(f"[{r['title']}] {r['href']}\n{r['body']}")
    except Exception as e:
        log.warning(f"Search failed: {e}")

    # Fetch existing Impressum page if available
    impressum_url = row.get("impressum_url") or row.get("website")
    if impressum_url:
        text = _fetch_page_text(impressum_url)
        if text:
            parts.append(f"\n=== Impressum page ({impressum_url}) ===\n{text}")

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Ollama
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = """Du bist ein Data-Cleaner-Agent für Pflegeheime in NRW.
Du erhältst bereits gescrapte Daten und frische Suchergebnisse aus dem Web.
Extrahiere die korrekten, verifizierten Kontaktdaten für genau diese Einrichtung.

Antworte AUSSCHLIESSLICH als valides JSON-Objekt mit diesen Feldern:
{
  "telefon": "...",
  "email": "...",
  "adresse": "...",
  "geschaeftsfuehrung": "...",
  "einrichtungsleitung": "...",
  "notes": "..."
}

Regeln:
- Nur Daten aus den vorliegenden Quellen — KEINE HALLUZINATIONEN
- Wenn keine verlässlichen Daten vorhanden: "No clear Data"
- Telefon im Format: 0XXX XXXXXXX oder +49 XXX XXXXXXX
- Email vollständig (user@domain.de)
- Adresse: Straße Nr, PLZ Ort
- notes: kurze Anmerkung falls Datenqualität unklar"""


def ollama_clean(row: dict, context: str) -> dict:
    user_msg = (
        f"Einrichtung: {row['name']}, {row['ort']} ({row.get('kreis','')})\n\n"
        f"{context}"
    )
    payload = {
        "model": OLLAMA_MODEL,
        "format": "json",
        "stream": False,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_msg},
        ],
        "options": {"temperature": 0.1},
    }
    content = ""
    try:
        resp = requests.post(
            f"{OLLAMA_HOST}/api/chat", json=payload, timeout=300
        )
        resp.raise_for_status()
        content = resp.json().get("message", {}).get("content", "") or ""
        return json.loads(content)
    except json.JSONDecodeError:
        log.warning("Ollama returned non-JSON — extracting from text")
        m = re.search(r"\{.*\}", content, re.DOTALL)
        try:
            return json.loads(m.group(0)) if m else {}
        except json.JSONDecodeError:
            log.error(f"Could not extract JSON from Ollama response: {content[:200]}")
            return {}
    except Exception as e:
        log.error(f"Ollama call failed: {e}")
        return {}


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------
XLSX_COLS = [
    ("api_id", "ID", 8),
    ("name", "Name", 42),
    ("traeger", "Träger", 16),
    ("ort", "Ort", 18),
    ("kreis", "Kreis", 22),
    ("telefon_clean", "Telefon", 22),
    ("email_clean", "Email", 32),
    ("email_status", "Email-Status", 12),
    ("adresse_clean", "Adresse", 45),
    ("geschaeftsfuehrung_clean", "Geschäftsführung", 28),
    ("einrichtungsleitung_clean", "Einrichtungsleitung", 28),
    ("gf_source", "GF-Quelle", 11),
    ("quality", "Qualität", 11),
    ("cleaner", "Cleaner", 11),
    ("clean_notes", "Notes", 50),
    ("website", "Website (Original)", 38),
    ("cleaned_at", "Cleaned At", 20),
]


def export_xlsx(conn, xlsx_path: str, cleaner_filter: str | None = None):
    """Write a formatted .xlsx (header styling, conditional formatting, summary sheet)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()
    ws = wb.active
    ws.title = "Pflegeheime Cleaned"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    data_font = Font(size=10)

    for i, (_, label, width) in enumerate(XLSX_COLS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    where = ["cleaned = TRUE"]
    params: list = []
    if cleaner_filter:
        where.append("cleaner = %s")
        params.append(cleaner_filter)
    sql = (
        f"SELECT {', '.join(c for c, _, _ in XLSX_COLS)} "
        f"FROM pflegeheime "
        f"WHERE {' AND '.join(where)} "
        f"ORDER BY ort, name"
    )

    with conn.cursor() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            if val is None:
                cell_val: object = ""
            elif hasattr(val, "tzinfo") and val.tzinfo is not None:
                cell_val = val.replace(tzinfo=None)
            else:
                cell_val = val
            cell = ws.cell(row=r_idx, column=c_idx, value=cell_val)
            cell.alignment = data_align
            cell.border = border
            cell.font = data_font

    if rows:
        qual_idx = next(i for i, (k, _, _) in enumerate(XLSX_COLS, start=1) if k == "quality")
        qual_col = get_column_letter(qual_idx)
        last = len(rows) + 1
        rng = f"{qual_col}2:{qual_col}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"ok"'],
            fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"suspect"'],
            fill=PatternFill("solid", fgColor="FFEB9C")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="equal", formula=['"empty"'],
            fill=PatternFill("solid", fgColor="FFC7CE")))
        ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Pflegeheime NRW — Cleaning Report"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = f"Filter: cleaner = {cleaner_filter or 'ALL'}"
    summary["A2"].font = Font(italic=True, size=10, color="666666")

    cleaner_idx = next(i for i, (k, _, _) in enumerate(XLSX_COLS) if k == "cleaner")
    quality_idx = next(i for i, (k, _, _) in enumerate(XLSX_COLS) if k == "quality")

    by_cleaner: dict[str, dict[str, int]] = {}
    for row in rows:
        c = row[cleaner_idx] or "unknown"
        q = row[quality_idx] or "unknown"
        by_cleaner.setdefault(c, {"ok": 0, "suspect": 0, "empty": 0, "total": 0})
        by_cleaner[c][q] = by_cleaner[c].get(q, 0) + 1
        by_cleaner[c]["total"] += 1

    summary["A4"] = "Cleaner"
    summary["B4"] = "OK"
    summary["C4"] = "Suspect"
    summary["D4"] = "Empty"
    summary["E4"] = "Total"
    for col in "ABCDE":
        summary[f"{col}4"].font = Font(bold=True)
        summary[f"{col}4"].fill = PatternFill("solid", fgColor="D9E1F2")

    r = 5
    for cleaner_name, stats in sorted(by_cleaner.items()):
        summary[f"A{r}"] = cleaner_name
        summary[f"B{r}"] = stats.get("ok", 0)
        summary[f"C{r}"] = stats.get("suspect", 0)
        summary[f"D{r}"] = stats.get("empty", 0)
        summary[f"E{r}"] = stats["total"]
        r += 1

    summary[f"A{r+1}"] = f"Exportiert: {len(rows)} Zeilen"
    summary[f"A{r+1}"].font = Font(italic=True, size=10)

    for col, w in [("A", 18), ("B", 10), ("C", 10), ("D", 10), ("E", 10)]:
        summary.column_dimensions[col].width = w

    wb.save(xlsx_path)
    log.info(f"Wrote xlsx: {xlsx_path}  ({len(rows)} rows)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def process_one(conn, cleaner: str = "ollama") -> bool:
    """Process one row. Returns True if a row was found, False if all done."""
    row = get_next_uncleaned(conn)
    if not row:
        log.info("All rows cleaned.")
        return False

    log.info(f"Cleaning: {row['name']} ({row['ort']})")

    context = gather_context(row)
    cleaned = ollama_clean(row, context)

    if not cleaned:
        cleaned = {k: NO_DATA for k in
                   ["telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung"]}
        cleaned["notes"] = "Ollama returned no data"

    cleaned = validate_cleaned(cleaned, row.get("ort", ""))
    save_cleaned(conn, row["api_id"], cleaned, cleaner=cleaner)
    log.info(
        f"[{cleaned.get('quality','?').upper():7}] {row['name']}: "
        f"Tel={cleaned.get('telefon','')}, "
        f"Email={cleaned.get('email','')}, "
        f"EL={cleaned.get('einrichtungsleitung','')}"
    )
    return True


def process_one_worker(cleaner: str) -> bool:
    """Worker variant: own DB connection, FOR UPDATE SKIP LOCKED, atomic commit."""
    conn = db_connect()
    conn.autocommit = False
    try:
        row = get_next_uncleaned(conn, for_update=True)
        if not row:
            conn.commit()
            return False

        log.info(f"Cleaning: {row['name']} ({row['ort']})")

        # Fast path: scraped fields already valid → skip search + LLM
        cleaned = heuristic_clean(row)
        used_cleaner = cleaner
        if cleaned is not None:
            used_cleaner = f"{cleaner}+heuristic"
        else:
            context = gather_context(row)
            cleaned = ollama_clean(row, context)
            if not cleaned:
                cleaned = {k: NO_DATA for k in
                           ["telefon", "email", "adresse", "geschaeftsfuehrung", "einrichtungsleitung"]}
                cleaned["notes"] = "Ollama returned no data"

        cleaned = validate_cleaned(cleaned, row.get("ort", ""))
        save_cleaned(conn, row["api_id"], cleaned, cleaner=used_cleaner)
        conn.commit()
        log.info(
            f"[{cleaned.get('quality','?').upper():7}] {row['name']} "
            f"({'H' if used_cleaner.endswith('+heuristic') else 'L'}): "
            f"Tel={cleaned.get('telefon','')}, "
            f"Email={cleaned.get('email','')}, "
            f"EL={cleaned.get('einrichtungsleitung','')}"
        )
        return True
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def main():
    global OLLAMA_MODEL
    parser = argparse.ArgumentParser(description="Pflegeheim Data Cleaner")
    parser.add_argument("--csv", default=CSV_PATH, help="Path to CSV file")
    parser.add_argument("--import-only", action="store_true", help="Only import CSV, skip cleaning")
    parser.add_argument("--loop", action="store_true", help="Run until all rows cleaned (or --limit)")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process at most N rows then stop (use with --loop)")
    parser.add_argument("--model", default=OLLAMA_MODEL, help="Ollama model")
    parser.add_argument("--cleaner-tag", default="ollama",
                        help="Label saved in pflegeheime.cleaner column (e.g. 'ollama' or 'langdock')")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay between rows in --loop mode")
    parser.add_argument("--xlsx", default=None,
                        help="After run: export cleaned rows to formatted .xlsx (with Summary sheet)")
    parser.add_argument("--xlsx-only", action="store_true",
                        help="Skip cleaning, just export current cleaned rows to --xlsx")
    parser.add_argument("--xlsx-cleaner", default=None,
                        help="Filter xlsx export by cleaner tag (e.g. 'ollama', 'langdock')")
    parser.add_argument("--workers", type=int, default=1,
                        help="Parallel worker threads (each with own DB conn + FOR UPDATE SKIP LOCKED). "
                             "Server-side: requires OLLAMA_NUM_PARALLEL >= workers for true concurrency.")
    args = parser.parse_args()

    OLLAMA_MODEL = args.model

    conn = db_connect()
    ensure_schema(conn)

    if args.xlsx_only:
        if not args.xlsx:
            parser.error("--xlsx-only requires --xlsx PATH")
        export_xlsx(conn, args.xlsx, cleaner_filter=args.xlsx_cleaner)
        conn.close()
        return

    if table_empty(conn):
        log.info("Table is empty — importing CSV first")
        import_csv(conn, args.csv)

    if args.import_only:
        conn.close()
        return

    processed = 0
    if args.loop and args.workers and args.workers > 1:
        # parallel worker pool
        import threading
        target = args.limit or 10**9
        lock = threading.Lock()
        stop = threading.Event()
        counter = {"n": 0}

        def worker_loop():
            while not stop.is_set():
                try:
                    got = process_one_worker(args.cleaner_tag)
                    if not got:
                        # No more rows in DB — really done
                        stop.set()
                        return
                    with lock:
                        counter["n"] += 1
                        if counter["n"] >= target:
                            log.info(f"Reached --limit {target}, stopping.")
                            stop.set()
                            return
                except Exception as exc:
                    # one-row failure must NOT kill the pool
                    log.error(f"Worker error (continuing): {exc}")
                if args.delay:
                    time.sleep(args.delay)

        threads = [threading.Thread(target=worker_loop, daemon=True)
                   for _ in range(args.workers)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        processed = counter["n"]
    elif args.loop:
        while process_one(conn, cleaner=args.cleaner_tag):
            processed += 1
            if args.limit and processed >= args.limit:
                log.info(f"Reached --limit {args.limit}, stopping.")
                break
            time.sleep(args.delay)
    else:
        if process_one(conn, cleaner=args.cleaner_tag):
            processed = 1

    log.info(f"Done. Processed {processed} rows this run.")

    if args.xlsx:
        export_xlsx(conn, args.xlsx, cleaner_filter=args.xlsx_cleaner)

    conn.close()


if __name__ == "__main__":
    main()
