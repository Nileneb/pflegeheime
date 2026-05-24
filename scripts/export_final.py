#!/usr/bin/env python3
"""Phase 3 export — the final, source-attributed list as .xlsx + .csv.

Reads the *_final columns produced by compose_final.py. Every contact field is
authoritative NRW open data; every Geschäftsführer is grounded in a fetched
Impressum (or honestly marked 'Nicht ermittelt'). The 'Quelle'/'GF-Quelle'
columns make the provenance auditable — that is what 'perfect' means here.

Usage:
  python scripts/export_final.py --xlsx out/pflegeheime_final.xlsx --csv data/pflegeheime_final.csv
"""
import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from data_cleaner import db_connect

COLS = [
    ("api_id", "ID", 8),
    ("name", "Name", 40),
    ("traeger_final", "Träger", 28),
    ("ort", "Ort", 16),
    ("kreis", "Kreis", 20),
    ("adresse_final", "Adresse", 42),
    ("telefon_final", "Telefon", 20),
    ("email_final", "E-Mail", 30),
    ("website_final", "Website", 32),
    ("geschaeftsfuehrung_final", "Geschäftsführung", 34),
    ("gf_source", "GF-Quelle", 30),
    ("einrichtungsleitung_final", "Einrichtungsleitung", 26),
    ("el_source", "EL-Quelle", 18),
    ("quality_final", "Qualität", 10),
    ("data_source_summary", "Quellen/Hinweise", 40),
]

SQL = f"SELECT {', '.join(c for c,_,_ in COLS)} FROM pflegeheime ORDER BY ort, name"


def export_csv(rows, path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow([label for _, label, _ in COLS])
        for r in rows:
            w.writerow(["" if v is None else v for v in r])
    print(f"CSV:  {path}  ({len(rows)} Zeilen)")


def export_xlsx(rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook(); ws = wb.active; ws.title = "Pflegeheime NRW"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1F4E78")
    ha = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
    da = Alignment(horizontal="left", vertical="top", wrap_text=True); df = Font(size=10)

    for i, (_, label, w) in enumerate(COLS, 1):
        c = ws.cell(1, i, label); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = bd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30; ws.freeze_panes = "A2"

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, "" if val is None else val)
            c.alignment = da; c.border = bd; c.font = df

    qi = next(i for i, (k, _, _) in enumerate(COLS, 1) if k == "quality_final")
    ql = get_column_letter(qi); last = len(rows) + 1; rng = f"{ql}2:{ql}{last}"
    for val, color in [("complete", "C6EFCE"), ("partial", "FFEB9C"),
                       ("check", "FFC7CE"), ("empty", "F2F2F2")]:
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal",
            formula=[f'"{val}"'], fill=PatternFill("solid", fgColor=color)))
    # highlight 'Nicht ermittelt' GF
    gi = next(i for i, (k, _, _) in enumerate(COLS, 1) if k == "geschaeftsfuehrung_final")
    gl = get_column_letter(gi)
    ws.conditional_formatting.add(f"{gl}2:{gl}{last}", CellIsRule(operator="equal",
        formula=['"Nicht ermittelt"'], fill=PatternFill("solid", fgColor="FCE4D6")))
    ws.auto_filter.ref = ws.dimensions

    # summary
    s = wb.create_sheet("Übersicht")
    s["A1"] = "Pflegeheime NRW — finale, quellenattribuierte Liste"
    s["A1"].font = Font(bold=True, size=14)
    conn = db_connect(); cur = conn.cursor()

    def one(q):
        cur.execute(q); return cur.fetchone()[0]
    N = one("SELECT count(*) FROM pflegeheime")
    rowsinfo = [
        ("Einrichtungen gesamt", N),
        ("mit Telefon (offiziell, API)", one("SELECT count(*) FROM pflegeheime WHERE telefon_final<>''")),
        ("mit E-Mail", one("SELECT count(*) FROM pflegeheime WHERE email_final<>''")),
        ("mit Website", one("SELECT count(*) FROM pflegeheime WHERE website_final<>''")),
        ("mit vollständiger Adresse", one("SELECT count(*) FROM pflegeheime WHERE adresse_final<>''")),
        ("mit Geschäftsführung (Impressum-belegt)",
         one("SELECT count(*) FROM pflegeheime WHERE geschaeftsfuehrung_final<>'' AND geschaeftsfuehrung_final<>'Nicht ermittelt'")),
        ("mit Einrichtungsleitung", one("SELECT count(*) FROM pflegeheime WHERE einrichtungsleitung_final<>'' AND einrichtungsleitung_final<>'Nicht ermittelt'")),
    ]
    r = 3
    for label, val in rowsinfo:
        s[f"A{r}"] = label; s[f"B{r}"] = val
        s[f"B{r}"] = f"{val}  ({100*val//N}%)"
        r += 1
    s["A11"] = "Qualität:"; s["A11"].font = Font(bold=True)
    cur.execute("SELECT quality_final, count(*) FROM pflegeheime GROUP BY quality_final ORDER BY 2 DESC")
    r = 12
    for q, n in cur.fetchall():
        s[f"A{r}"] = q; s[f"B{r}"] = n; r += 1
    s.column_dimensions["A"].width = 40; s.column_dimensions["B"].width = 18
    conn.close()

    wb.save(path)
    print(f"XLSX: {path}  ({len(rows)} Zeilen)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="out/pflegeheime_final.xlsx")
    ap.add_argument("--csv", default="data/pflegeheime_final.csv")
    args = ap.parse_args()
    conn = db_connect(); cur = conn.cursor()
    cur.execute(SQL); rows = cur.fetchall(); conn.close()
    os.makedirs(os.path.dirname(args.xlsx) or ".", exist_ok=True)
    export_csv(rows, args.csv)
    export_xlsx(rows, args.xlsx)


if __name__ == "__main__":
    main()
